import schedule
import time
import numpy as np
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import os
import shutil
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import win32com.client as win32
import pandas as pd
from datetime import datetime, timedelta
import datetime as dt
import schedule
import time
import numpy as np
import xlrd
import urllib.request
import requests
import sys
import getpass
import json
import schedule
import time
import lxml
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

#znajduje ostatnią uzupełnioną datę w pliku
wb = load_workbook(filename="abc.xlsx")
ws = wb["a"]
ostatni_wiersz=ws.max_row
ostatnia_data=ws.cell(row=ostatni_wiersz,column=1).value  #uchwycona ostatnia data, dla której są dane w pliku excel

holidays = ["2023-04-10","2023-05-01","2023-05-03","2023-06-08","2023-08-15","2023-11-01","2023-11-11","2023-12-25","2023-12-26","2024-01-01"]

# kod  na ostatni dzień roboczy
dzisiaj=dt.date.today()
delta1=dt.timedelta(days=1)
delta2=dt.timedelta(days=2)
ostatni_dzien=dzisiaj-delta1

if ostatni_dzien.weekday() == 5:
    ostatni_dzien = ostatni_dzien - delta1
elif ostatni_dzien.weekday() == 6:
    ostatni_dzien = ostatni_dzien - delta2

ostatni_dzien_str=str(ostatni_dzien)

for x in range(len(holidays)):
    if ostatni_dzien_str in holidays:
        ostatni_dzien = ostatni_dzien - delta1
        if ostatni_dzien.weekday() == 5:
            ostatni_dzien = ostatni_dzien - delta1
        elif ostatni_dzien.weekday() == 6:
            ostatni_dzien = ostatni_dzien - delta2
        ostatni_dzien_str = str(ostatni_dzien)


weekdays = [5,6]
data_poczatkowa=dt.datetime.strptime(ostatnia_data,"%d-%m-%Y")+delta1     #trzeba do ostatniej daty dodać 1 dzień
start_day = data_poczatkowa
end_day = ostatni_dzien


daterange = pd.date_range(start_day, end_day)
for date in daterange:
    if date.weekday() not in weekdays and date.strftime("%Y-%m-%d") not in holidays:
        dzien = date.strftime("%d-%m-%Y")

        sciezkaWebDriver=r"C:\Users\psliwa\PycharmProjects\Pobieranie_danych_tge\chromedriver.exe"  # do ściezki doklejam chromedriver, który wczesniej instaluje ze strony (sprawdź wersje chrome i sciągnij odpowiedni chromedriver)
        # https://chromedriver.chromium.org/downloads     link do sciągniecia chromedrivera
        #ileDni=1
        driver = webdriver.Chrome(executable_path=sciezkaWebDriver)
        url = 'https://tge.pl/energia-elektryczna-otf?dateShow='+dzien+'&dateAction=prev'
        page = driver.get(url)
        time.sleep(1)
        #-----BASE ---
        df = pd.read_html(driver.page_source, header = 0, decimal=",", thousands='.')
        df[0] = df[0].drop('Unnamed: 1',axis=1)  #df[0] dla base   df[1] dla peak
        base = df[0]
        base = base.iloc[:-1]         #usunięcie ostatniego wiersza z tabeli (podsumowania)
        base.insert(0,'data',dzien)    #dodanie daty w pierwszej kolumnie(0-zerowej)
        base['typ'] = "BASE"            #dodajemy kolumne Typ: "BASE" dla produktów z tabeli base, PEAK dla produktów z tabeli PEAK
        #-----PEAK-----
        df[1] = df[1].drop('Unnamed: 1',axis=1)
        peak = df[1]
        peak = peak.iloc[:-1]
        peak.insert(0, 'data', dzien)
        peak['typ'] = 'PEAK'

        wb = load_workbook(filename="abc.xlsx")
        ws = wb["a"]
        for x in dataframe_to_rows(base, index=False, header=False):
            ws.append(x)          #append dodaje dane do już istniejących w pliku
        wb.save("abc.xlsx")
        for x in dataframe_to_rows(peak, index=False, header=False):
            ws.append(x)
        wb.save("abc.xlsx")

        driver.close()
        driver.quit()



